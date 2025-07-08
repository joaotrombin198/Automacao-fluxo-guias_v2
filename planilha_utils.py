# planilha_utils.py

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font

fill_ok = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
fill_erro = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
font_arial_normal = Font(name='Arial', bold=False)
font_arial_bold = Font(name='Arial', bold=True)

def carregar_planilha(caminho):
    df = pd.read_excel(caminho, dtype={"NR_SEQ_SEGURADO": str})
    df["NR_SEQ_SEGURADO"] = df["NR_SEQ_SEGURADO"].str.strip()
    return df

def salvar_planilha_formatada(df, caminho_saida):
    print(f"[INFO] Salvando planilha atualizada em {caminho_saida} com formatação...")
    df.to_excel(caminho_saida, index=False)
    wb = openpyxl.load_workbook(caminho_saida)
    ws = wb.active

    # STATUS
    status_col = None
    for cell in ws[1]:
        if cell.value and str(cell.value).upper() == "STATUS":
            status_col = cell.column
            break
    if status_col:
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=status_col)
            valor = (cell.value or "").lower()
            if any(p in valor for p in ["erro", "falha", "inválido", "excluído"]):
                cell.fill = fill_erro
                cell.font = font_arial_bold
            elif valor.strip() != "":
                cell.fill = fill_ok
                cell.font = font_arial_normal
            else:
                cell.fill = PatternFill(fill_type=None)
                cell.font = font_arial_normal

    # NR_SEQ_SEGURADO como texto
    carteira_col = None
    for cell in ws[1]:
        if cell.value and str(cell.value).upper() == "NR_SEQ_SEGURADO":
            carteira_col = cell.column
            break
    if carteira_col:
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=carteira_col)
            cell.number_format = '@'
            cell.font = font_arial_normal

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            if cell.font.name != 'Arial' or cell.font.bold:
                cell.font = font_arial_normal

    wb.save(caminho_saida)
    print("[INFO] Planilha salva e formatada com sucesso.")
